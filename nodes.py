# -*- coding: utf-8 -*-

import openpyxl

variety_book = openpyxl.load_workbook('variety.xlsx')
mechanism_book = openpyxl.load_workbook('mechanism.xlsx')
pg_book = openpyxl.load_workbook('species.xlsx')


def variety_nodes():
    variety = variety_book.sheetnames
    variety.pop(0)
    return variety


def mechanism_nodes():
    mechanism = mechanism_book.sheetnames
    mechanism.pop(0)
    return mechanism


def phylum_node():
    phylum = pg_book.sheetnames
    return phylum[3:]


def genus_node():
    sheet = pg_book['network']
    genus_cells = sheet['C']
    genus = list()
    for i in genus_cells[1:]:
        genus.append(i.value)
    return genus


def phylum_group(name):
    data = pg_book[name]['A']
    genus = list()
    for i in data:
        genus.append(i.value)
    return genus


def phylum_dict():
    group = dict()
    for phylum in phylum_node():
        group[phylum] = phylum_group(phylum)
    return group


def sample_nodes():
    sample = 'CK-F, CK-T, CK-S, HU-F, HU-T, HU-S, TH-F, TH-T, TH-S, TT-F, TT-T, TT-S'
    return sample.split(', ')


def logic_nodes():
    root = 'main'
    main = ['time', 'concentration']

    return root, main


def assay_node():
    assay_data = variety_book[variety_book.sheetnames[0]]['A']
    assay = list()
    for i in assay_data:
        assay.append(i.value)
    assay.pop(0)
    return assay


def in_group(name__, group__, book__):
    if name__ not in group__:
        raise ValueError('%s not in workbook' % name__)
    else:
        sheet = book__[name__]
        in_assay = sheet['A']
        result = list()
        for i in in_assay:
            result.append(i.value)
        result.pop(0)
    return result


def in_variety(name):
    all_variety = variety_book.sheetnames
    book_v = variety_book
    return in_group(name__=name, group__=all_variety, book__=book_v)


def variety_group():
    all_ = variety_nodes()
    group = dict()
    for var in all_:
        group[var] = in_variety(var)
    return group


def mechanism_group():
    all_ = mechanism_nodes()
    group = dict()
    for mec in all_:
        group[mec] = in_mechanism(mec)
    return group


def in_mechanism(name):
    all_mechanism = mechanism_book.sheetnames
    book_m = mechanism_book
    return in_group(name__=name, group__=all_mechanism, book__=book_m)


def cluster(G):
    flag = 0
    variety = variety_group()
    for k in variety.keys():
        G.add_nodes_from(variety[k], flag=flag)
        flag += 1
    phylum = phylum_dict()
    phylum.pop('Others')
    for kk in phylum.keys():
        G.add_nodes_from(phylum[kk], flag=flag)
        flag += 1


if __name__ == '__main__':
    x = variety_group()
    print(x)
