# -*- coding: utf-8 -*-

import networkx as nx
import openpyxl as px
import nodes as nd

# from network4 import row_map, column_map


data = px.load_workbook("xzx.xlsx")
data_ = px.load_workbook("显著性.xlsx")
sheet = data['Sheet1']
sheet_ = data_[data_.sheetnames[0]]

r_book = px.load_workbook('相关性2.xlsx')
xgx = r_book[r_book.sheetnames[0]]


def get_nodes():
    nodes = list()
    for i in sheet['A']:
        nodes.append(i.value)
    nodes.pop(0)
    return nodes


def nodes_map():
    nodes = get_nodes()
    nodes_map__ = dict()
    for i in enumerate(nodes):
        nodes_map__[i[1]] = i[0] + 1
    return nodes_map__


def get_cell_value(row, column):
    return sheet_.cell(row, column).value


def get_r_value(row, column):
    return xgx.cell(row, column).value


def draw():
    G = nx.Graph()
    nodes = get_nodes()
    G.add_nodes_from(nodes)
    for i in nodes:
        for j in nodes:
            value = get_cell_value(nodes_map()[i], nodes_map()[j])
            if value > 0.01:
                weight = 0
            # elif value > 0.001 and value <= 0.01:
            #     weight = abs(xgx.cell(row_map()[i], column_map()[j]).value)
            else:
                weight = get_r_value(nodes_map()[i], nodes_map()[j])

            G.add_edge(i, j, weight=weight)
    nd.cluster(G)
    nx.write_gml(G, 'C:\\users\\administrator\\desktop\\network_test.gml')


def main():
    draw()


if __name__ == '__main__':
    main()
