# -*- coding: utf-8 -*-


import openpyxl as px
import networkx as nx
import nodes as nd

# from network3 import cluster

book = px.load_workbook('xd.xlsx')
name_sheet = book[book.sheetnames[0]]
data_sheet = book[book.sheetnames[1]]

r_book = px.load_workbook('相关性2.xlsx')
xgx = r_book[r_book.sheetnames[0]]

genus = list()
genes = list()

for i in name_sheet['A']:
    if i.value:
        genus.append(i.value)

for j in name_sheet[1]:
    if j.value:
        genes.append(j.value)


def row_map():
    row_dict = dict()
    for i_ in enumerate(genus):
        row_dict[i_[1]] = i_[0] + 1
    return row_dict


def column_map():
    column_dict = dict()
    for k in enumerate(genes):
        column_dict[k[1]] = k[0] + 1
    return column_dict


def draw():
    genus_nodes = genus
    genes_nodes = genes
    row = row_map()
    column = column_map()
    G = nx.DiGraph()
    G.add_nodes_from(genus)
    G.add_nodes_from(genes)
    for e1 in genus_nodes:
        for e2 in genes_nodes:
            value = data_sheet.cell(row[e1], column[e2]).value
            if value > 0.01:
                weight = 0
            else:
                weight = 500000000 * abs(xgx.cell(row[e1], column[e2]).value)

            if xgx.cell(row[e1], column[e2]).value < 0:
                edge_color_flag = -1
            elif xgx.cell(row[e1], column[e2]).value == 0:
                edge_color_flag = 0
            else:
                edge_color_flag = 1
            # elif value > 0.001 and value <= 0.01:
            #     weight = 1  # * xgx.cell(row[e1], column[e2]).value
            # else:
            #     weight = 1  # * abs(xgx.cell(row[e1], column[e2]).value)
            G.add_edge(e1, e2, weight=weight, capacity=edge_color_flag)
    nd.cluster(G)
    nx.write_gml(G, 'C:\\users\\administrator\\desktop\\network_test4.gml')


def main():
    draw()


if __name__ == '__main__':
    main()
